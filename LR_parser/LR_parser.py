token_stream=open("LA_output.txt",'r')          # Load output file of scanner

line_stream=open("TokenLine.txt","r")



scanner_output = token_stream.readline()

line_stream_output=line_stream.readline()

stream_token = scanner_output.split()           # lexical analyzer output

each_token = line_stream_output.split()         #
stream_token.append('$')





for i in range(stream_token.count('INT')):      # The code Matching token of Scanner and token of LR_parser
    index=stream_token.index('INT')
    stream_token[index]='vtype'
# INT = vtype
for i in range(stream_token.count('CHAR')):
    index = stream_token.index('CHAR')
    stream_token[index] = 'vtype'
# CHAR = vtype
for i in range(stream_token.count('IF')):
    index=stream_token.index('IF')
    stream_token[index]='if'
# IF = if
for i in range(stream_token.count('ELSE')):
    index=stream_token.index('ELSE')
    stream_token[index]='else'
# ELSE = else
for i in range(stream_token.count('WHILE')):
    index=stream_token.index('WHILE')
    stream_token[index]='while'
# WHILE = while
for i in range(stream_token.count('RETURN')):
    index=stream_token.index('RETURN')
    stream_token[index]='return'
# RETURN = return
for i in range(stream_token.count('ZERO')):
    index=stream_token.index('ZERO')
    stream_token[index]='num'
# ZERO = num
for i in range(stream_token.count('NZERO')):
    index=stream_token.index('NZERO')
    stream_token[index]='num'
# NZERO =  num
for i in range(stream_token.count('ADDSUB')):
    index=stream_token.index('ADDSUB')
    stream_token[index]='addsub'
# ADDSUB = addsub
for i in range(stream_token.count('MULTDIV')):
    index=stream_token.index('MULTDIV')
    stream_token[index]='multdiv'
# MULTDIV = multdiv
for i in range(stream_token.count('SEMICOLON')):
    index=stream_token.index('SEMICOLON')
    stream_token[index]='semi'
# SEMICOLON = semi
for i in range(stream_token.count('LBRACE')):
    index=stream_token.index('LBRACE')
    stream_token[index]='lbrace'
# LBRACE = lbrace
for i in range(stream_token.count('RBRACE')):
    index=stream_token.index('RBRACE')
    stream_token[index]='rbrace'
# RBRACE = rbrace
for i in range(stream_token.count('LPAREN')):
    index=stream_token.index('LPAREN')
    stream_token[index]='lparen'
# LPAREN = lparen
for i in range(stream_token.count('RPAREN')):
    index=stream_token.index('RPAREN')
    stream_token[index]='rparen'
# RPAREN = rparen
for i in range(stream_token.count('COMMA')):
    index=stream_token.index('COMMA')
    stream_token[index]='comma'
# COMMA = comma
for i in range(stream_token.count('STRING')):
    index=stream_token.index('STRING')
    stream_token[index]='literal'
# STRING = literal
for i in range(stream_token.count('COMPARISON')):
    index=stream_token.index('COMPARISON')
    stream_token[index]='comp'
# COMPARISON = comp
for i in range(stream_token.count('ASSIGNMENT')):
    index=stream_token.index('ASSIGNMENT')
    stream_token[index]='assign'
# ASSIGNMENT = assign
for i in range(stream_token.count('IDENTIFIER')):
    index=stream_token.index('IDENTIFIER')
    stream_token[index]='id'
# IDENTIFIER = id
for i in range(stream_token.count('INTEGER')):
    index=stream_token.index('INTEGER')
    stream_token[index]='num'
# INTEGER = num




stack_list = ['00']          # Stack implemented as list

Rlist = stream_token         # Llist | Rlist

Llist = []

current_state="00"
                                            # production_rule_left
production_rule_left = {"00":"CODE'",\
                        "01":"CODE",\
                        "02":"CODE",\
                        "03":"VDECL",\
                        "04":"FDECL",\
                        "05":"ARG",\
                        "06":"ARG",\
                        "07":"MOREARGS",\
                        "08":"MOREARGS",\
                        "09":"BLOCK",\
                        "10":"BLOCK",\
                        "11":"BLOCK",\
                        "12":"STMT",\
                        "13":"STMT",\
                        "14":"STMT",\
                        "15":"STMT",\
                        "16":"STMT",\
                        "17":"RHS",\
                        "18":"RHS",\
                        "19":"RHS",\
                        "20":"EXPR",\
                        "21":"EXPR",\
                        "22":"TERM",\
                        "23":"TERM",\
                        "24":"FACTOR",\
                        "25":"FACTOR",\
                        "26":"FACTOR",\
                        "27":"FCALL",\
                        "28":"COND",\
                        "29":"RETURN",\
                        "30":"CODE",\
                        "31":"CODE"\
                        }

                                                        # production_rule_right
production_rule_right = {"00":["CODE"], \
                         "01":["VDECL","CODE"],\
                         "02":["FDECL","CODE"],\
                         "03":["vtype","id","semi"],\
                         "04":["vtype","id", "lparen","ARG","lbrace","BLOCK"],\
                         "05":["vtype","id","MOREARGS"],\
                         "06":["rparen"],\
                         "07":["comma","vtype","id","MOREARGS"],\
                         "08":["rparen"],\
                         "09":["STMT","BLOCK"],\
                         "10":["rbrace"],\
                         "11":["RETURN","rbrace"],\
                         "12":["VDECL"],\
                         "13":["id","assign","RHS","semi"],\
                         "14":["if","lparen","COND","rparen","lbrace","BLOCK","else","lbrace","BLOCK"],\
                         "15":["while","lparen","COND","rparen","lbrace","BLOCK"],\
                         "16":["FCALL","semi"],\
                         "17":["EXPR"],\
                         "18":["FCALL"],\
                         "19":["literal"],\
                         "20":["TERM","addsub","EXPR"],\
                         "21":["TERM"],\
                         "22":["FACTOR","multdiv","TERM"],\
                         "23":["FACTOR"],\
                         "24":["lparen","EXPR","rparen"],\
                         "25":["id"],\
                         "26":["num"],\
                         "27":["id","lparen","ARG"],\
                         "28":["FACTOR","comp","FACTOR"],\
                         "29":["return","FACTOR","semi"],\
                         "30":["VDECL"],\
                         "31":["FDECL"]\
                         }



# Functions defined #####################

# Function to determine if it is a number


def isNum(string):
    try:
        float(string)
        return True

    except ValueError:
        return False




# Function to get (row, col) from stream

def row_col(Llist,Rlist,current_state):

    row=current_state
    try:
        col=Rlist[0]
    except IndexError:

        return row, col

    return row,col


# Shift function

def shift(SLR_table_output, Llist,Rlist,stack_list,current_state):

    stack_list.append(SLR_table_output[-2:])

    current_state=SLR_table_output[-2:]

    Rlist.reverse()
    Llist.append(Rlist.pop())
    Rlist.reverse()

    return Llist,Rlist,stack_list,current_state

# error function

def error(count,each_token):
    print("error")
    print("There is no Action or Goto in SLR table\n","error occur",each_token[count],"line")
    exit()




# GOTO function

def GOTO(SLR_table_output, Llist,Rlist,stack_list,current_state):

    stack_list.append(SLR_table_output)
    current_state=SLR_table_output

    return Llist,Rlist,stack_list,current_state


# SLR_table function
# key of dictionary is (state, terminal or non terminal)
# value of dictionary is shift or reduce or error or accept (ex:S02,R05,..)
import openpyxl

oneNum = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']

SLR_dic = {}

wb = openpyxl.load_workbook('DFA, SLR table.xlsx')
sheet = wb['Sheet1']
A1 = sheet.cell(row=1, column=1)


for i in range(4, 79):
    for j in range(2, 37):
        cell = sheet.cell(row=i, column=j)
        value = str(cell.value)
        if value != 'None':
            if (value[0] == 's' or value[0] == 'r') and len(value) == 2 :
                value = value[0].upper() + '0' + value[1]
            if (value[0] == 's' or value[0] == 'r') and len(value) == 3 :
                value = value[0].upper() + value[1] + value[2]

            if value in oneNum:
                value = '0' + value
            key1 = str(sheet.cell(i, 1).value)
            if key1 in oneNum:
                key1 = '0' + key1
            key2 = str(sheet.cell(3, j).value)
            if key2 == 'RETURN2':
                key2 = 'RETURN'
            SLR_dic[(key1, key2)] = value
#for i in SLR_dic.keys():
    #print(i, SLR_dic[i])

def SLR_table(row, col, SLR_dic):
    return SLR_dic[(row, col)]

# reduce function

def reduce(SLR_table_output, Llist,Rlist,stack_list,production_rule_left,production_rule_right,current_state):

    PR_LEFT = production_rule_left[SLR_table_output[-2:]]

    PR_RIGHT = production_rule_right[SLR_table_output[-2:]]


    for i in range(len(PR_RIGHT)):

        Llist.pop()

        stack_list.pop()

    Llist.append(PR_LEFT)

    return Llist,Rlist,stack_list,current_state




# distinguish function (reduce, shift, accept, error, GOTO)

def distinguish(SLR_table_output, Llist,Rlist,stack_list,production_rule_left,production_rule_right,current_state):

    if SLR_table_output[0]=="R":
        return reduce(SLR_table_output, Llist,Rlist,stack_list,production_rule_left,production_rule_right,current_state)

    elif SLR_table_output[0]=="S":
        return shift(SLR_table_output, Llist,Rlist,stack_list,current_state)

    elif isNum(SLR_table_output[0]):
        return GOTO(SLR_table_output, Llist,Rlist,stack_list,current_state)

    elif SLR_table_output[0]=="A":
        print("accept")
        exit()
    else :
        error(count, each_token)


row,col=row_col(Llist,Rlist, current_state)  # initialize row, col

SLR_table_output=SLR_table(row,col, SLR_dic)    # initialize SLR table output

# print(SLR_table_output)

pre_state=[] # log list for states

count=0

while True:
    '''
    print("count",count,"token")
    pre_state .append(current_state)

    print("현재 상태 :",current_state)
    print("상태 로그 :", pre_state)
    print("현재 Llist :", Llist,"| Rlist :",Rlist)
    print("토큰 값 :",col)
    print("현재 스택 :",stack_list)
    print("SLR_table_output 값 :",SLR_table_output,"\n")
    '''
    Llist, Rlist, stack_list, current_state = distinguish(SLR_table_output, Llist,Rlist,stack_list,production_rule_left,production_rule_right,current_state)  # initialize values
    if SLR_table_output[0] == "S":
        count+=1

    if SLR_table_output[0] == "R":

        temp=current_state
        current_state=stack_list[-1]
        Rlist.reverse()
        Rlist.append(Llist[-1])
        Rlist.reverse()
        row, col = row_col(Llist,Rlist, current_state)
        Rlist.reverse()
        Rlist.pop()
        Rlist.reverse()

    else:

        row, col = row_col(Llist,Rlist, current_state)

    try:

        SLR_table_output = SLR_table(row, col,SLR_dic)

    except KeyError:
        '''
        print("\nKeyError 발생\n\n(",row,",",col,") 키 값에 해당하는 SLR_table 의 dic value 존재하지 않음\n")
        print("현재 상태 :", current_state)
        print("상태 로그 :", pre_state)
        print("현재 Llist :", Llist, "| Rlist :", Rlist)
        print("토큰 값 :", col)
        print("현재 스택 :", stack_list)
        print("SLR_table_output 값 :", SLR_table_output, "\n")
        '''

        error(count,each_token)